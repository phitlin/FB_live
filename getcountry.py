import openpyxl
from urllib2 import urlopen
from json import load 
import json, re, os, sys, requests


openfile = open('key.txt')
key = openfile.read()
print key

##source file should be called latslongs.xlsx and lat/long should be columns 1 and 2

wb = openpyxl.load_workbook('latslongs.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

rownum = 1


def getval():
	for rowOfCellObjects in sheet['A'+ str(rownum):'B'+ str(rownum)]:
		for cellObj in rowOfCellObjects:
			print(cellObj.value)

		lat = sheet['A' + str(rownum)].value
		lng = sheet['B' + str(rownum)].value

		a = str(lat)
		b = str(lng)

		url = 'https://maps.googleapis.com/maps/api/geocode/json?latlng=' + a +',' + b + '&components=country&key=' + key
		print url		

		response = urlopen(url)
		json_obj = load(response)
		w = json_obj
		print w["status"]

		if w["status"] == 'OK':
			x = 0
			y = 0
			print y
			while y != 'country':
				print x
				print 'not yet'
				x = x + 1
				y = w["results"][0]["address_components"][x]["types"][0]
				print y
			print w["results"][0]["address_components"][x]["long_name"]
			country = w["results"][0]["address_components"][x]["long_name"]
			sheet.cell(row=rownum, column=5).value = country
			sheet.cell(row=rownum, column=6).value = url
		else:
			sheet.cell(row=rownum, column=5).value = 'cannot find'	

		wb.save('with country.xlsx')



			
## set to number of rows

while rownum < 35:
	rownum = rownum + 1
	getval()


